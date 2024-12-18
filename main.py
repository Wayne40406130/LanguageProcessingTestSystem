import datetime
import json
import os
import random
import time
import tkinter as tk
from tkinter import messagebox

import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows


class LanguageProcessingTestSystem:
    def __init__(
        self,
        root,
        font_size=16,
        font_family="Microsoft JhengHei",
        stage_order=["formal", "reward", "penalty", "reward_penalty"],
        config_path=r"C:\Users\USER\Desktop\LanguageProcessingTestSystem-main\LanguageProcessingTestSystem-main\words_config.json",
    ):
        self.root = root
        self.root.title("詞彙判斷試驗系統")
        self.root.attributes("-fullscreen", True)  # 設置全屏顯示
        self.root.bind("<Escape>", self.exit_fullscreen)  # 綁定 Escape 鍵退出全屏

        self.participant_name = ""
        self.group = ""
        self.accuracy_threshold = 0.8
        self.stage_order = stage_order
        self.font = (font_family, font_size)  # 使用指定字體

        self.words_config = self.load_words_from_config(config_path)["types"]
        if not self.words_config:
            raise ValueError("Failed to load words configuration")

        self.correct_answers = 0
        self.current_question_count = 0  # 當前已回答的問題數
        self.timeout_id = None
        self.current_stage_index = 0
        self.current_balance = 0  # 初始化金額

        # 計數器
        self.true_word_count = 0
        self.false_word_count = 0
        self.pm_target_count = 0
        self.true_word_correct = 0
        self.false_word_correct = 0
        self.pm_target_correct = 0

        # 答對率
        self.true_word_accuracy = 0
        self.false_word_accuracy = 0
        self.pm_target_accuracy = 0

        # 變量來跟踪當前的階段
        self.current_stage = ""

        self.summary_data = {
            "time": [datetime.datetime.now().strftime("%Y-%m-%d_%Hh%M")],
            "practice": [],
            "lexical_prac": [],
            "keyresponse_prac": [],
            "lexical_ans_prac": [],
            "lexical_crate_prac": [],
            "phonetic_ans_prac": [],
            "phonetic_crate_prac": [],
            "reactiontime_prac": [],
            "reactiontime_avg_prac": [],
        }

        for stage in ["nofb", "rfb", "pfb", "rpfb"]:
            self.summary_data[f"lexical_{stage}"] = []
            self.summary_data[f"keyresponse_{stage}"] = []
            self.summary_data[f"lexical_ans_{stage}"] = []
            self.summary_data[f"lexical_crate_{stage}"] = []
            self.summary_data[f"phonetic_ans_{stage}"] = []
            self.summary_data[f"phonetic_crate_{stage}"] = []
            self.summary_data[f"reactiontime_{stage}"] = []
            self.summary_data[f"reactiontime_avg_{stage}"] = []
            if stage in ["rfb", "pfb", "rpfb"]:  # 金錢變化相關的階段
                self.summary_data[f"accum_{stage}"] = []  # 新增累積金額欄位

        self.results_data = {
            "practice": [],
            "formal": [],
            "reward": [],
            "penalty": [],
            "reward_penalty": [],
        }

        # # 單詞清單，按指定順序排好
        # self.word_list = self.create_word_list()

        # GUI設置
        self.setup_gui()

    def load_words_from_config(self, config_path):
        """根據JSON配置檔載入所有階段的詞彙"""
        required_stages = [
            "practice",
            "formal",
            "reward",
            "penalty",
            "reward_penalty",
        ]  # 必須包含的五個階段

        if not os.path.exists(config_path):
            messagebox.showerror("錯誤", f"Config file {config_path} 不存在。")
            return None

        try:
            with open(config_path, "r", encoding="utf-8") as file:
                config_data = json.load(file)  # 載入JSON檔案

            missing_stages = [
                stage for stage in required_stages if stage not in config_data["types"]
            ]

            if missing_stages:
                missing_stages_str = ", ".join(missing_stages)
                messagebox.showerror(
                    "錯誤", f"缺少以下階段的詞彙配置: {missing_stages_str}"
                )
                return None

            return config_data  # 如果所有必須的階段都存在，回傳完整的配置

        except json.JSONDecodeError:
            messagebox.showerror("錯誤", "JSON 文件格式錯誤，無法解析。")
            return None

        except Exception as e:
            messagebox.showerror("錯誤", f"讀取配置文件時發生錯誤: {e}")
            return None

    def select_words_for_stage(self, stage):
        """根據當前階段選擇對應的詞彙，支持動態鍵值"""
        try:
            # 取得對應階段的詞彙區塊
            stage_words = self.words_config[stage]

            word_types = list(stage_words.keys())

            # 使用鍵名動態分配對應的詞彙
            self.true_word_type = word_types[0]  # 動態取得真詞類型
            self.false_word_type = word_types[1]  # 動態取得假詞類型
            self.pm_target_type = word_types[2]  # 動態取得 PM target 類型

            # 將詞彙類型的內容分配到變量
            self.true_words = set(stage_words[self.true_word_type])
            self.false_words = set(stage_words[self.false_word_type])
            self.pm_targets = stage_words[self.pm_target_type]

            print(f"Loaded words for stage: {stage}")
        except KeyError as e:
            print(f"Error: Stage '{stage}' not found in words_config - {e}")
            raise ValueError(f"Missing words configuration for stage: {stage}")

    def create_word_list(self):
        total_words = (
            len(self.true_words) + len(self.false_words) + len(self.pm_targets)
        )
        word_list = [None] * total_words  # 初始化詞彙列表

        # 將 PM target 移到指定的位置，並檢查是否超出範圍
        for target, pos in self.pm_targets.items():
            if pos - 1 >= total_words or pos - 1 < 0:
                messagebox.showerror(
                    "錯誤",
                    f"詞彙總長度為: {total_words}。PM target 「{target}」 的位置 「{pos}」 超出範圍，請重新調整 words_config.json。",
                )
                self.root.destroy()  # 結束程序
                return  # 確保程式中斷
            word_list[pos - 1] = (target, self.pm_target_type)

        # 將其餘的詞隨機填入空位
        remaining_words = [(word, self.true_word_type) for word in self.true_words] + [
            (word, self.false_word_type) for word in self.false_words
        ]

        random.shuffle(remaining_words)

        # 將未指定順序的詞彙填充到空位
        for i in range(len(word_list)):
            if word_list[i] is None and remaining_words:
                word_list[i] = remaining_words.pop(0)

        return word_list

    def exit_fullscreen(self, event=None):
        """退出全屏"""
        self.root.attributes("-fullscreen", False)

    def setup_gui(self):
        """設置初始界面"""
        self.root.configure(bg="black")  # 設置背景為黑色

        self.instructions_label = tk.Label(
            self.root,
            text="測試",
            font=self.font,
            fg="white",
            bg="black",
        )
        self.instructions_label.pack()

        self.name_label = tk.Label(
            self.root, text="參與者姓名：", font=self.font, fg="white", bg="black"
        )
        self.name_label.pack()
        self.name_entry = tk.Entry(
            self.root, font=self.font, fg="white", bg="black", insertbackground="white"
        )
        self.name_entry.pack()

        self.group_label = tk.Label(
            self.root, text="組別：", font=self.font, fg="white", bg="black"
        )
        self.group_label.pack()
        self.group_entry = tk.Entry(
            self.root, font=self.font, fg="white", bg="black", insertbackground="white"
        )
        self.group_entry.pack()

        self.start_button = tk.Button(
            self.root,
            text="開始",
            command=self.start_experiment,
            font=self.font,
            fg="white",
            bg="black",
        )
        self.start_button.pack()

        # 顯示金額的Label
        self.balance_label = tk.Label(
            self.root, text="", font=self.font, fg="white", bg="black"
        )
        self.balance_label.pack(anchor="nw", padx=10, pady=10)

    def update_balance_label(self):
        """更新金額顯示"""
        self.balance_label.config(text=f"金額: {self.current_balance} 元")
        self.balance_label.pack(anchor="nw", padx=10, pady=10)  # 確保顯示在左上角

    def start_experiment(self):
        """開始實驗"""
        self.participant_name = self.name_entry.get()
        self.group = self.group_entry.get()

        if not self.participant_name or not self.group:
            messagebox.showerror("錯誤", "請輸入姓名和組別。")
            return

        self.run_practice_instructions()

    def run_practice_instructions(self):
        """顯示練習指導語"""
        print(f"run_practice_instructions")
        # 清除所有現有的窗口內容
        for widget in self.root.winfo_children():
            widget.pack_forget()

        self.title_label = tk.Label(
            self.root,
            text="練習階段",
            font=(self.font[0], self.font[1] + 30),  # 字體加大
            fg="white",
            bg="black",
        )
        self.title_label.pack(side="top", pady=50)  # 放置在頂部，並增加一些間距

        # 顯示前導詞
        self.instructions_label = tk.Label(
            self.root,
            text="真實的詞彙請按「A」，非真實的詞彙請按「L」，\n"
            "但當詞彙中的注音包含「ㄍ」或「ㄐ」時，請按「空白鍵」。\n",
            font=self.font,
            fg="white",
            bg="black",
        )
        self.instructions_label.pack(expand=True)

        # 綁定Enter鍵事件以開始練習
        self.root.bind("<Return>", self.show_fixed_words)

    def show_fixed_words(self, event=None):
        """顯示固定字詞，等待按鍵響應"""
        self.root.unbind("<Return>")
        self.fixed_words = ["民提", "橘子", "歌曲"]
        self.fixed_word_index = 0  # 追踪目前顯示的字詞索引

        # 開始顯示第一個固定字詞
        self.show_next_fixed_word()

    def show_next_fixed_word(self):
        """依序顯示固定字詞"""
        if self.fixed_word_index < len(self.fixed_words):
            # 清空畫面
            for widget in self.root.winfo_children():
                widget.pack_forget()

            # 顯示固定字詞
            self.instructions_label = tk.Label(
                self.root,
                text=self.fixed_words[self.fixed_word_index],
                font=self.font,
                fg="white",
                bg="black",
            )
            self.instructions_label.pack(expand=True)

            # 綁定按鍵響應
            self.root.bind("<Key>", self.handle_fixed_word_response)
        else:
            # 所有固定字詞顯示完成後，顯示 "按任意鍵開始"
            self.show_any_key_screen(None)

    def handle_fixed_word_response(self, event):
        """處理固定字詞的按鍵響應"""
        key = event.keysym.lower()
        allowed_keys = ["a", "l", "space"]
        if key in allowed_keys:
            # 如果按鍵有效，繼續顯示下一個固定字詞
            self.fixed_word_index += 1
            self.show_next_fixed_word()
        else:
            print(f"Ignored key: {key}, not in allowed keys: {allowed_keys}")

    def show_any_key_screen(self, event):
        """顯示按任意鍵開始屏幕"""
        self.root.unbind("<Return>")
        for widget in self.root.winfo_children():
            widget.pack_forget()

        self.instructions_label = tk.Label(
            self.root, text="按任意鍵開始", font=self.font, fg="white", bg="black"
        )
        self.instructions_label.pack(expand=True)

        self.root.bind("<Key>", self.start_practice)

    def start_practice(self, event):
        """開始練習"""
        self.root.unbind("<Key>")
        self.current_stage = "practice"
        self.select_words_for_stage(self.current_stage)
        self.run_practice()

    def run_practice(self):
        """運行練習"""
        self.word_list = self.create_word_list()  # 重置詞彙列表
        self.show_black_screen_before_next_word(stage="practice")

    def show_black_screen_before_next_word(self, stage):
        """顯示全黑屏幕500ms，然後顯示下一個單詞"""
        print(f"show_black_screen_before_next_word:{stage}")
        self.show_black_screen()
        self.root.after(500, lambda: self.show_next_word(stage))

    def show_black_screen(self):
        """顯示全黑屏幕"""
        for widget in self.root.winfo_children():
            widget.pack_forget()

        self.root.configure(bg="black")
        self.root.update()

    def show_next_word(self, stage):
        """顯示下一個單詞"""
        print(f"show_next_word:{stage}")
        for widget in self.root.winfo_children():
            widget.pack_forget()

        if stage == "reward" or stage == "penalty" or stage == "reward_penalty":
            self.update_balance_label()  # 更新金額顯示

        if self.word_list:
            self.current_word, self.current_key = self.word_list.pop(0)
            self.start_time = time.time()  # 記錄開始顯示單詞的時間
            self.instructions_label.config(
                text=self.current_word, font=self.font, fg="white", bg="black"
            )
            self.instructions_label.pack(expand=True)
            self.root.bind("<Key>", lambda event: self.check_answer(event, stage))
            self.timeout_id = self.root.after(
                3000, lambda: self.check_answer_timeout(stage)
            )
        else:
            self.end_stage(stage)

    def check_answer(self, event, stage):
        """檢查答案"""
        print(f"check_answer stage:{stage}")
        key = event.keysym.lower()

        # 設置有效鍵
        allowed_keys = ["a", "l", "space"]

        # 無效按鍵處理：如果按下的鍵不在有效鍵列表中，則不處理
        if key not in allowed_keys:
            print(f"Ignored key: {key}, not in allowed keys: {allowed_keys}")
            # 確保無效鍵不會打斷流程，繼續綁定按鍵事件
            self.root.bind("<Key>", lambda event: self.check_answer(event, stage))
            return  # 只忽略該次輸入，並等待下一次有效按鍵

        # 有效按鍵處理
        print(f"check_answer stage:{stage}")
        if self.timeout_id is not None:
            self.root.after_cancel(self.timeout_id)
            self.timeout_id = None

        # 解除鍵盤綁定，只在有效按鍵時解除
        self.root.unbind("<Key>")
        reaction_time = int(
            (time.time() - self.start_time) * 1000
        )  # 將反應時間從秒轉為毫秒

        # 保存反應時間和按鍵響應到results_data中
        self.results_data[stage].append(
            {
                "word": self.current_word,
                "response": key,
                "reaction_time": reaction_time,
                "correct_response": self.true_word_type
                if self.current_word in self.true_words
                else (
                    self.false_word_type
                    if self.current_word in self.false_words
                    else self.pm_target_type
                ),
            }
        )

        if self.current_word in self.true_words:
            self.true_word_count += 1
            if key == self.true_word_type:
                self.true_word_correct += 1
            self.true_word_accuracy = self.true_word_correct / self.true_word_count
            print(f"self.true_word_count: {self.true_word_count}")
            print(f"self.true_word_correct: {self.true_word_correct}")
            print(f"True word accuracy: {self.true_word_accuracy:.2%}")
        elif self.current_word in self.false_words:
            self.false_word_count += 1
            if key == self.false_word_type:
                self.false_word_correct += 1
            self.false_word_accuracy = self.false_word_correct / self.false_word_count
            print(f"self.false_word_count: {self.false_word_count}")
            print(f"self.false_word_correct: {self.false_word_correct}")
            print(f"False word accuracy: {self.false_word_accuracy:.2%}")
        elif self.current_word in self.pm_targets:
            self.pm_target_count += 1
            if key == self.pm_target_type:
                self.pm_target_correct += 1
                if stage in ["reward", "reward_penalty"]:
                    self.reward_user()  # 在獎勵或獎懲階段獎勵用戶
                    return
            else:
                if stage in ["penalty", "reward_penalty"]:
                    self.penalize_user()  # 在懲罰或獎懲階段處罰用戶
                    return
            print(f"self.pm_target_count: {self.pm_target_count}")
            print(f"self.pm_target_correct: {self.pm_target_correct}")
            self.pm_target_accuracy = self.pm_target_correct / self.pm_target_count
            print(f"PM target accuracy: {self.pm_target_accuracy:.2%}")
        stage_prefix = self.get_stage_prefix(stage)
        if stage_prefix in ["rfb", "pfb", "rpfb"]:
            # 將當前金額追加到相應的summary_data欄位
            self.summary_data[f"accum_{stage_prefix}"].append(self.current_balance)

        self.show_black_screen_before_next_word(stage)

    def reward_user(self):
        """獎勵用戶"""
        self.current_balance += 10
        stage_prefix = self.get_stage_prefix(self.current_stage)  # 獲取當前階段的前綴

        if stage_prefix in ["rfb", "rpfb"]:  # 在獎勵或獎懲階段更新金額
            self.summary_data[f"accum_{stage_prefix}"].append(self.current_balance)

        # 呼叫顯示獎勵信息函數，傳入正確的當前階段
        self.show_reward_message(stage=self.current_stage)

    def show_reward_message(self, stage):
        """顯示獎勵信息"""
        print(f"show_reward_message for stage: {stage}")
        self.show_black_screen()
        self.instructions_label = tk.Label(
            self.root,
            text=f"獲得十元\n目前金額: {self.current_balance}元",
            font=self.font,
            fg="white",
            bg="black",
        )
        self.instructions_label.pack(expand=True)
        self.root.update()
        self.root.after(
            1500,
            lambda: self.update_balance_and_continue(
                stage=stage
            ),  # 使用傳遞的 stage 參數
        )

    def penalize_user(self):
        """懲罰用戶"""
        print(f'"""懲罰用戶"""')
        self.current_balance -= 10
        stage_prefix = self.get_stage_prefix(self.current_stage)  # 獲取當前階段的前綴
        print(f"stage_prefix:{stage_prefix}")

        if stage_prefix in ["pfb", "rpfb"]:  # 在懲罰或獎懲階段更新金額
            print("# 在懲罰或獎懲階段更新金額")
            self.summary_data[f"accum_{stage_prefix}"].append(self.current_balance)
            print(self.summary_data[f"accum_{stage_prefix}"])

        # 呼叫顯示懲罰信息函數，傳入正確的當前階段
        self.show_penalty_message(stage=self.current_stage)

    def show_penalty_message(self, stage):
        """顯示懲罰信息"""
        print(f"show_penalty_message for stage: {stage}")
        self.show_black_screen()
        self.instructions_label = tk.Label(
            self.root,
            text=f"扣除十元\n目前金額: {self.current_balance}元",
            font=self.font,
            fg="white",
            bg="black",
        )
        self.instructions_label.pack(expand=True)
        self.root.update()
        self.root.after(
            1500,
            lambda: self.update_balance_and_continue(
                stage=stage
            ),  # 使用傳遞的 stage 參數
        )

    def update_balance_and_continue(self, stage):
        """更新金額並繼續"""
        self.update_balance_label()
        self.show_black_screen_before_next_word(stage)

    def check_answer_timeout(self, stage):
        """超時檢查答案"""
        print(f"check_answer_timeout:{stage}")
        if self.timeout_id is not None:
            self.timeout_id = None
            self.root.unbind("<Key>")
        reaction_time = 3000  # 固定為3000毫秒
        key = ""  # 沒有按鍵響應

        # 保存超時反應到results_data中
        self.results_data[stage].append(
            {
                "word": self.current_word,
                "response": key,
                "reaction_time": reaction_time,
                "correct_response": self.true_word_type
                if self.current_word in self.true_words
                else (
                    self.false_word_type
                    if self.current_word in self.false_words
                    else self.pm_target_type
                ),
            }
        )

        if self.current_word in self.true_words:
            self.true_word_count += 1
            self.true_word_accuracy = self.true_word_correct / self.true_word_count
            print(f"self.true_word_count: {self.true_word_count}")
            print(f"self.true_word_correct: {self.true_word_correct}")
            print(f"True word accuracy: {self.true_word_accuracy:.2%}")
        elif self.current_word in self.false_words:
            self.false_word_count += 1
            self.false_word_accuracy = self.false_word_correct / self.false_word_count
            print(f"self.false_word_count: {self.false_word_count}")
            print(f"self.false_word_correct: {self.false_word_correct}")
            print(f"False word accuracy: {self.false_word_accuracy:.2%}")
        elif self.current_word in self.pm_targets:
            self.pm_target_count += 1
            if stage == "penalty" or stage == "reward_penalty":
                print(f"在懲罰或獎懲階段才執行扣錢邏輯")
                self.penalize_user()
                return

            self.pm_target_accuracy = self.pm_target_correct / self.pm_target_count
        print(f"self.pm_target_count: {self.pm_target_count}")
        print(f"self.pm_target_correct: {self.pm_target_correct}")
        print(f"PM target accuracy: {self.pm_target_accuracy:.2%}")

        # 在超時檢查答案後，即使沒有金額變動，也更新金額到 summary_data
        stage_prefix = self.get_stage_prefix(stage)
        if stage_prefix in ["rfb", "pfb", "rpfb"]:
            # 將當前金額追加到相應的summary_data欄位
            self.summary_data[f"accum_{stage_prefix}"].append(self.current_balance)

        self.show_black_screen_before_next_word(stage)

    def end_stage(self, stage):
        """結束階段"""
        if stage == "practice":
            self.current_stage = "practice"
            self.end_practice()
        elif stage == "formal":
            self.current_stage = "formal"
            self.end_formal_stage()
        elif stage == "reward":
            self.current_stage = "reward"
            self.end_reward_stage()
        elif stage == "penalty":
            self.current_stage = "penalty"
            self.end_penalty_stage()
        elif stage == "reward_penalty":
            self.current_stage = "reward_penalty"
            self.end_reward_penalty_stage()

    def end_practice(self):
        """結束練習"""
        true_word_accuracy = (
            self.true_word_correct / self.true_word_count if self.true_word_count else 0
        )
        false_word_accuracy = (
            self.false_word_correct / self.false_word_count
            if self.false_word_count
            else 0
        )

        # 設定當前階段為 'practice'
        self.current_stage = "practice"

        # 保存當前階段的結果，會逐次添加而不覆蓋
        self.save_stage_results()

        # 根據正確率決定是否進入下一個階段
        if (
            true_word_accuracy < self.accuracy_threshold
            or false_word_accuracy < self.accuracy_threshold
        ):
            self.reset_counters()
            self.run_practice_instructions()
        else:
            self.start_next_stage()

    def start_next_stage(self, event=None):
        """開始下一階段"""
        self.root.unbind("<Key>")
        self.run_main_experiment()

    def run_main_experiment(self):
        """運行主要實驗"""
        if self.current_stage_index < len(self.stage_order):
            stage = self.stage_order[self.current_stage_index]
            self.current_stage_index += 1
            if stage == "formal":
                print("formal")
                self.current_stage = "formal"
                self.show_instructions(
                    "formal",
                    "真實的詞彙請按「A」，非真實的詞彙請按「L」\n"
                    "但當詞彙中的注音含有「ㄅ」或「ㄉ」時，請按「空白鍵」。\n",
                )
            elif stage == "reward":
                print("reward")
                self.current_stage = "reward"
                self.show_instructions(
                    "reward",
                    "真實的詞彙請按「A」，非真實的詞彙請按「L」\n"
                    "但當詞彙中的注音含有「ㄆ」或「ㄊ」時，請按「空白鍵」。\n"
                    "\n"
                    "每當您正確辨認出含有「ㄆ」與「ㄊ」的詞彙時，\n"
                    "會顯示您獲得10元，且會累計顯示於左上角，\n"
                    "若您了解此實驗的程序請按enter鍵開始。",
                )
            elif stage == "penalty":
                print("penalty")
                self.current_stage = "penalty"
                self.show_instructions(
                    "penalty",
                    "請判斷螢幕上的詞彙是否為真實存在的詞彙，\n"
                    "真實的詞彙請按「A」，非真實的詞彙請按「L」\n"
                    "但當詞彙中的注音含有「ㄇ」或「ㄋ」時，請按「空白鍵」。\n"
                    "\n"
                    "每當您未正確辨認出含有「ㄇ」與「ㄋ」的詞彙時，\n"
                    "會顯示您被扣除10元，且會累計顯示於左上角，\n"
                    "若您了解此實驗的程序請按enter鍵開始。",
                )
            elif stage == "reward_penalty":
                print("reward_penalty")
                self.current_stage = "reward_penalty"
                self.show_instructions(
                    "reward_penalty",
                    "請判斷螢幕上的詞彙是否為真實存在的詞彙，\n"
                    "真實的詞彙請按「A」，非真實的詞彙請按「L」\n"
                    "但當詞彙中的注音含有「ㄈ」或「ㄌ」時，請按「空白鍵」。\n"
                    "\n"
                    "每當您正確辨認出含有「ㄈ」的詞彙時，\n"
                    "會顯示您獲得10元，且會累計顯示於左上角，\n"
                    "每當您未正確辨認出含有「ㄌ」的詞彙時，\n"
                    "會顯示您被扣除10元，且會累計顯示於左上角，\n"
                    "若您了解此實驗的程序請按enter鍵開始。",
                )
        else:
            self.save_results()
            self.show_thank_you_message()

    def show_instructions(self, stage, instructions):
        """顯示每個階段的指導語"""
        self.instructions_label.config(
            text=instructions, font=self.font, fg="white", bg="black"
        )
        self.instructions_label.pack(expand=True)
        self.root.bind(
            "<Return>", lambda event: self.show_any_key_screen_next(event, stage)
        )

    def show_any_key_screen_next(self, event, stage):
        """顯示按任意鍵開始屏幕"""
        self.root.unbind("<Return>")
        for widget in self.root.winfo_children():
            widget.pack_forget()

        self.instructions_label = tk.Label(
            self.root, text="按任意鍵開始", font=self.font, fg="white", bg="black"
        )
        self.instructions_label.pack(expand=True)

        self.root.bind("<Key>", lambda event: self.start_stage(event, stage))

    def start_stage(self, event, stage):
        """開始階段"""
        print(f"stage{stage}")
        self.select_words_for_stage(stage)
        self.root.unbind("<Key>")
        self.word_list = self.create_word_list()  # 重置詞彙列表
        if stage == "formal":
            self.run_formal_stage()
        elif stage == "reward":
            self.run_reward_stage()
        elif stage == "penalty":
            self.run_penalty_stage()
        elif stage == "reward_penalty":
            self.run_reward_penalty_stage()

    def run_formal_stage(self):
        """運行正式測試階段"""
        self.reset_counters()
        self.select_words_for_stage(self.current_stage)
        self.show_black_screen_before_next_word(stage="formal")

    def end_formal_stage(self):
        """結束正式測試階段"""
        self.current_stage = "formal"
        self.save_stage_results()
        self.start_next_stage()

    def run_reward_stage(self):
        """運行獎勵階段"""
        self.reset_counters()
        self.current_balance = 200  # 初始金額
        self.select_words_for_stage(self.current_stage)
        self.update_balance_label()  # 顯示金額
        self.show_black_screen_before_next_word(stage="reward")

    def end_reward_stage(self):
        """結束獎勵階段"""
        self.current_stage = "reward"
        self.save_stage_results()
        self.clear_balance_label()  # 清除金額顯示
        self.start_next_stage()

    def run_penalty_stage(self):
        """運行懲罰階段"""
        print("run_penalty_stage")
        self.reset_counters()
        self.current_balance = 200  # 初始金額
        self.select_words_for_stage(self.current_stage)
        self.update_balance_label()  # 顯示金額
        self.show_black_screen_before_next_word(stage="penalty")

    def end_penalty_stage(self):
        """結束懲罰階段"""
        print("end_penalty_stage")
        self.current_stage = "penalty"
        print(f"self.current_stage:{self.current_stage}")
        self.save_stage_results()
        self.clear_balance_label()  # 清除金額顯示
        self.start_next_stage()

    def run_reward_penalty_stage(self):
        """運行獎懲階段"""
        self.reset_counters()
        self.current_balance = 200  # 初始金額
        self.select_words_for_stage(self.current_stage)
        self.update_balance_label()  # 顯示金額
        self.show_black_screen_before_next_word(stage="reward_penalty")

    def end_reward_penalty_stage(self):
        """結束獎懲階段"""
        self.current_stage = "reward_penalty"
        self.save_stage_results()
        self.clear_balance_label()  # 清除金額顯示
        self.start_next_stage()

    def clear_balance_label(self):
        """清除金額顯示"""
        self.balance_label.config(text="")

    def reset_counters(self):
        """重置計數器"""
        self.correct_answers = 0
        self.current_question_count = 0
        self.true_word_count = 0
        self.false_word_count = 0
        self.pm_target_count = 0
        self.true_word_correct = 0
        self.false_word_correct = 0
        self.pm_target_correct = 0

    def save_stage_results(self):
        """保存當前階段的結果到 summary_data"""
        print(f"Saving results for stage: {self.current_stage}")
        stage_prefix = self.get_stage_prefix(self.current_stage)

        # 取得當前階段的單詞列表
        current_results = self.results_data[self.current_stage]

        # 將每一個詞語結果保存到對應的 summary_data 欄位中
        for result in current_results:
            self.summary_data[f"lexical_{stage_prefix}"].append(result["word"])
            self.summary_data[f"keyresponse_{stage_prefix}"].append(result["response"])

            if result["correct_response"] == self.pm_target_type:
                self.summary_data[f"phonetic_ans_{stage_prefix}"].append(
                    result["correct_response"]
                )
                self.summary_data[f"lexical_ans_{stage_prefix}"].append("")  # 對應空值
            else:
                self.summary_data[f"lexical_ans_{stage_prefix}"].append(
                    result["correct_response"]
                )
                self.summary_data[f"phonetic_ans_{stage_prefix}"].append("")  # 對應空值

            self.summary_data[f"reactiontime_{stage_prefix}"].append(
                result["reaction_time"]
            )

        # 檢查是否是金錢變化階段，並記錄金額變化
        if stage_prefix in ["rfb", "pfb", "rpfb"]:
            if f"accum_{stage_prefix}" not in self.summary_data:
                self.summary_data[f"accum_{stage_prefix}"] = []  # 初始化金額變化列表
            self.summary_data[f"accum_{stage_prefix}"] += [""] * (
                len(self.summary_data[f"lexical_{stage_prefix}"])
                - len(self.summary_data[f"accum_{stage_prefix}"])
            )
        # 計算當前階段的正確率並保存
        lexical_accuracy = self.calculate_lexical_accuracy()
        phonetic_accuracy = self.calculate_phonetic_accuracy()

        self.summary_data[f"lexical_crate_{stage_prefix}"].append(
            round(lexical_accuracy, 2)
        )
        self.summary_data[f"phonetic_crate_{stage_prefix}"].append(
            round(phonetic_accuracy, 2)
        )

        # 計算平均反應時間並保存
        reaction_times = [result["reaction_time"] for result in current_results]
        average_reaction_time = (
            int(sum(reaction_times) / len(reaction_times)) if reaction_times else 0
        )
        self.summary_data[f"reactiontime_avg_{stage_prefix}"].append(
            round(average_reaction_time, 2)
        )
        print("!!!!!!!!")
        print(f"save_stage_results:")
        print(self.summary_data)
        # 清空當前階段的結果
        self.results_data[self.current_stage] = []

    def get_stage_prefix(self, stage):
        """根據當前的階段返回對應的前綴"""
        if stage == "practice":
            return "prac"
        elif stage == "formal":
            return "nofb"
        elif stage == "reward":
            return "rfb"
        elif stage == "penalty":
            return "pfb"
        elif stage == "reward_penalty":
            return "rpfb"
        else:
            raise ValueError(f"未知的階段: {stage}")

    def calculate_lexical_accuracy(self):
        """計算真詞和假詞的總正確率"""
        total_words = self.true_word_count + self.false_word_count
        if total_words == 0:
            return 0
        correct_words = self.true_word_correct + self.false_word_correct
        return (correct_words / total_words) * 100

    def calculate_phonetic_accuracy(self):
        """計算PM target的正確率"""
        if self.pm_target_count == 0:
            return 0
        return (self.pm_target_correct / self.pm_target_count) * 100

    def save_results(self):
        """保存結果到Excel文件"""
        filename = f"{self.group}.xlsx"
        if not os.path.exists(filename):
            workbook = openpyxl.Workbook()
            # 如果是新文件，先移除默認創建的工作表
            workbook.remove(workbook.active)
        else:
            # 文件存在，則打開該文件
            workbook = openpyxl.load_workbook(filename)

        # 如果已有以 participant_name 命名的工作表，刪除該工作表以覆蓋
        if self.participant_name in workbook.sheetnames:
            del workbook[self.participant_name]

        # 創建一個新的工作表
        worksheet = workbook.create_sheet(title=self.participant_name)

        # 處理每個階段的數據
        for stage in ["prac", "nofb", "rfb", "pfb", "rpfb"]:
            print("===============")
            print(f"stage:{stage}")
            # 根據階段名稱動態生成對應的欄位名稱
            lexical_key = f"lexical_{stage}"
            keyresponse_key = f"keyresponse_{stage}"
            lexical_ans_key = f"lexical_ans_{stage}"
            lexical_crate_key = f"lexical_crate_{stage}"
            phonetic_ans_key = f"phonetic_ans_{stage}"
            phonetic_crate_key = f"phonetic_crate_{stage}"
            reactiontime_key = f"reactiontime_{stage}"
            reactiontime_avg_key = f"reactiontime_avg_{stage}"
            accum_key = f"accum_{stage}" if stage in ["rfb", "pfb", "rpfb"] else None

            # 獲取對應欄位的數據
            if self.summary_data[lexical_key]:
                max_len = max(
                    len(self.summary_data[lexical_key]),
                    len(self.summary_data[keyresponse_key]),
                    len(self.summary_data[lexical_ans_key]),
                    len(self.summary_data[lexical_crate_key]),
                    len(self.summary_data[phonetic_ans_key]),
                    len(self.summary_data[phonetic_crate_key]),
                    len(self.summary_data[reactiontime_key]),
                    len(self.summary_data[reactiontime_avg_key]),
                    len(self.summary_data[accum_key])
                    if accum_key
                    else 0,  # 新增金錢變化的欄位長度
                )

                # 將所有列表填充到相同長度
                time_list = [self.summary_data["time"][0]] + [""] * (max_len - 1)
                practice_list = [""] * max_len

                self.summary_data[lexical_key] += [""] * (
                    max_len - len(self.summary_data[lexical_key])
                )
                self.summary_data[keyresponse_key] += [""] * (
                    max_len - len(self.summary_data[keyresponse_key])
                )
                self.summary_data[lexical_ans_key] += [""] * (
                    max_len - len(self.summary_data[lexical_ans_key])
                )
                self.summary_data[lexical_crate_key] += [""] * (
                    max_len - len(self.summary_data[lexical_crate_key])
                )
                self.summary_data[phonetic_ans_key] += [""] * (
                    max_len - len(self.summary_data[phonetic_ans_key])
                )
                self.summary_data[phonetic_crate_key] += [""] * (
                    max_len - len(self.summary_data[phonetic_crate_key])
                )
                self.summary_data[reactiontime_key] += [""] * (
                    max_len - len(self.summary_data[reactiontime_key])
                )
                self.summary_data[reactiontime_avg_key] += [""] * (
                    max_len - len(self.summary_data[reactiontime_avg_key])
                )
                if accum_key:  # 如果有累積金額的欄位，填充到相同長度
                    self.summary_data[accum_key] += [""] * (
                        max_len - len(self.summary_data[accum_key])
                    )
                # 構建 summary_data 字典
                summary_data = {
                    "time": time_list,
                    f"{stage}": practice_list,
                    lexical_key: self.summary_data[lexical_key],
                    keyresponse_key: self.summary_data[keyresponse_key],
                    lexical_ans_key: self.summary_data[lexical_ans_key],
                    lexical_crate_key: self.summary_data[lexical_crate_key],
                    phonetic_ans_key: self.summary_data[phonetic_ans_key],
                    phonetic_crate_key: self.summary_data[phonetic_crate_key],
                    reactiontime_key: self.summary_data[reactiontime_key],
                    reactiontime_avg_key: self.summary_data[reactiontime_avg_key],
                }
                print(f"summary_data")
                print(summary_data)
                # 新增金錢變化的欄位
                print("accum_key")
                print(accum_key)
                if accum_key:
                    summary_data[accum_key] = self.summary_data[accum_key]
                summary_df = pd.DataFrame(summary_data)

                for r in dataframe_to_rows(summary_df, index=False, header=True):
                    worksheet.append(r)

            # 保存 Excel 文件
            workbook.save(filename)

    def show_thank_you_message(self):
        """顯示銘謝詞並停留"""
        for widget in self.root.winfo_children():
            widget.pack_forget()

        self.instructions_label = tk.Label(
            self.root,
            text="感謝您的參與，測驗已完成。",
            font=self.font,
            fg="white",
            bg="black",
        )
        self.instructions_label.pack(expand=True)
        # 不綁定任何事件，停留在此屏幕


if __name__ == "__main__":
    root = tk.Tk()
    app = LanguageProcessingTestSystem(
        root,
        font_size=32,
        font_family="Microsoft JhengHei",
        stage_order=[
            "penalty",
            "reward",
            "reward_penalty",
            "formal",
        ],
    )
    root.mainloop()
